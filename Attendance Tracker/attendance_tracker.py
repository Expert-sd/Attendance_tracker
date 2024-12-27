import openpyxl
from collections import defaultdict



#STEP1. Load and read the attendance data
wb = openpyxl.load_workbook('attendance_sheet.xlsx')
sheet = wb['Attendance']
wb.active


#Create an empty dictionary to store the data


attendance_report = defaultdict(int)

#STEP2. Loop through each row to identify the employees present for each day and add it to the attendance_report dictionary.

for row_num in range(2, sheet.max_row + 1):
    employee_name = sheet.cell(row=row_num, column=3).value
    attendance_status= sheet.cell(row=row_num, column=4).value

    #print(f"Row {row_num}: Employee: {employee_name}, Status: {attendance_status}")

    if employee_name and attendance_status and attendance_status.strip().lower() == 'present':
    
       attendance_report[employee_name] += 1

    #print(f"Updated report: {employee_name} -> {attendance_report[employee_name]}")



#STEP3. Write the results to a new excel file
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = 'Employee Attendance'

#Add headers
new_sheet.append(['Employee Name', 'Number of days present'])

#Write data to the new excel file.
for employee_name, attendance_status in attendance_report.items():
    new_sheet.append([employee_name, attendance_status])

print(f"Writing to sheet: {employee_name} -> {attendance_status}")

new_wb.save('employee_attendance_report.xlsx')

print('Attendance report saved to "employee_attendance_report.xlsx".')